import openpyxl
from pathlib import Path
from datetime import datetime
import pandas as pd

xls = Path('.', 'MTR.202104.xlsx')
wb = openpyxl.load_workbook(xls)
print(wb.sheetnames)
# ['汇总', 'MTR使用手机记录', '手机同步记录', '每天拜访门店数', '门店拜访记录']

sht3 = wb['手机同步记录']
sht5 = wb['门店拜访记录']


def ptime(s):
    return datetime.strptime(s, '%Y-%m-%d %H:%M:%S.%f')


def ptime1(s):
    return datetime.strptime(s, '%Y-%m-%d %H:%M:%S')


lst3 = []
dct3 = {}
for i, row in enumerate(sht3.iter_rows(values_only=True)):
    if i == 0:
        continue  # skip first row
    code = row[2]
    name = row[3]
    start = row[5]
    start_time = ptime(start)
    device = row[9]
    it = (code, name, start_time, device)
    lst3.append(it)
    if code in dct3.keys():
        xs = dct3[code]
    else:
        xs = []
    xs.append(it)
    dct3[code] = xs

    # break

# print(lst3)

# check code, start_time in lst, find recenct device


def recentDevice(lst, code, start_time):
    msec = 60 * 60 * 24  # one day
    mdev, mtime = None, None
    # for code1, name1, start_time1, device1 in lst:
    for code1, name1, start_time1, device1 in dct3[code]:
        if code1 == code:
            dt = start_time - start_time1
            sec = dt.total_seconds()
            sec = sec if sec > 0 else -sec
            if sec < msec:
                msec, mdev, mtime = sec, device1, start_time1
    if msec <= 60 * 60:
        return (mdev, mtime)
    return (None, mtime)


lst5 = []
for i, row in enumerate(sht5.iter_rows(values_only=True)):
    if i == 0:
        continue
    trans = row[0]
    code = row[4]
    name = row[5]
    start, end = row[7], row[8]
    start_time, end_time = ptime1(start), ptime1(end)
    device, login_time = recentDevice(lst3, code, start_time)
    client_name, client_code = row[1], row[2]
    lst5.append((trans, client_name, client_code, code, name,
                 start_time, end_time, login_time, device))
    # if len(lst5) >= 100:
    #     break

df = pd.DataFrame(lst5, columns=['trans', 'client_name', 'client_code',
                                 'code', 'name', 'start_time', 'end_time', 'login_time', 'device'])
print(df)
df.to_excel('yaya.xlsx')
